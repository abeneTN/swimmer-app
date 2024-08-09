from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import PatternFill


print("Loading files...")

# Load the HTML content from the specified file path
file_path_2024_1 = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT D'HIVER M_C-J_S - 08_02_2024 ¤ 10_02_2024 - RADES.html"
file_path_2024_2 = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT DE TUNISIE D'ÉTÉ M_C J_S - 29_07_2024 ¤ 03_08_2024 - RADES.html"
file_path_2023 = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT M_C J_S - 14_08_2023 ¤ 17_08_2023 - RADES.html"
file_path_2024_benjamins_hiver = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT D'HIVER BENJAMINS - 05_02_2024 ¤ 07_02_2024 - RADES.html"
file_path_2024_benjamins_ete = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT D'ÉTÉ DE TUNISIE BENJAMINS - 25_07_2024 ¤ 27_07_2024 - RADES.html"
file_path_2023_benjamins_ete = "/home/abene/work/club-africain-analysis/files/CHAMPIONNAT DE TUNISIE BENJAMINS - 27_07_2023 ¤ 29_07_2023 - RADES.html"

file_paths = [file_path_2024_1, file_path_2024_2, file_path_2023, file_path_2023_benjamins_ete, file_path_2024_benjamins_hiver, file_path_2024_benjamins_ete]

with open(file_path_2024_1, 'r', encoding='windows-1252') as file:
    html_content = file.read()

print("Parsing HTML content...")

# Parse the HTML content
soup = BeautifulSoup(html_content, 'html.parser')
soup2 = BeautifulSoup(open(file_path_2024_2, 'r', encoding='windows-1252').read(), 'html.parser')
soup3 = BeautifulSoup(open(file_path_2023, 'r', encoding='windows-1252').read(), 'html.parser')
soup4 = BeautifulSoup(open(file_path_2024_benjamins_hiver, 'r', encoding='windows-1252').read(), 'html.parser')
soup5 = BeautifulSoup(open(file_path_2024_benjamins_ete, 'r', encoding='windows-1252').read(), 'html.parser')
soup6 = BeautifulSoup(open(file_path_2023_benjamins_ete, 'r', encoding='windows-1252').read(), 'html.parser')

soups = [soup, soup2, soup3, soup4, soup5, soup6]

debug_mode = ""

while debug_mode == "":
    print("3. Enable debug mode")
    debug_mode = input("Do you want to enable debug mode? (type ""yes"" / [hit enter] for no): ").lower() == 'yes'
    
# Initialize a dictionary to store the extracted data
current_event = None

event_names = {
    "01": "50 m NAGE LIBRE DAMES",
    "51": "50 m NAGE LIBRE MESSIEURS",
    "02": "100 m NAGE LIBRE DAMES",
    "52": "100 m NAGE LIBRE MESSIEURS",
    "03": "200 m NAGE LIBRE DAMES",
    "53": "200 m NAGE LIBRE MESSIEURS",
    "04": "400 m NAGE LIBRE DAMES",
    "54": "400 m NAGE LIBRE MESSIEURS",
    "05": "800 m NAGE LIBRE DAMES",
    "55": "800 m NAGE LIBRE MESSIEURS",
    "06": "1500 m NAGE LIBRE DAMES",
    "56": "1500 m NAGE LIBRE MESSIEURS",
    "43": "4 x 100 m NAGE LIBRE DAMES",
    "93": "4 x 100 m NAGE LIBRE MESSIEURS", 
    "44": "4 x 200 m NAGE LIBRE DAMES",
    "94": "4 x 200 m NAGE LIBRE MESSIEURS",
    "11": "50 m DOS DAMES",
    "61": "50 m DOS MESSIEURS",
    "12": "100 m DOS DAMES",
    "62": "100 m DOS MESSIEURS",
    "13": "200 m DOS DAMES",
    "63": "200 m DOS MESSIEURS",
    "21": "50 m BRASSE DAMES",
    "71": "50 m BRASSE MESSIEURS",
    "22": "100 m BRASSE DAMES",
    "72": "100 m BRASSE MESSIEURS",
    "23": "200 m BRASSE DAMES",
    "73": "200 m BRASSE MESSIEURS",
    "31": "50 m PAPILLON DAMES",
    "81": "50 m PAPILLON MESSIEURS",
    "32": "100 m PAPILLON DAMES",
    "82": "100 m PAPILLON MESSIEURS",
    "33": "200 m PAPILLON DAMES",
    "83": "200 m PAPILLON MESSIEURS",
    "41": "200 m 4 NAGES DAMES",
    "91": "200 m 4 NAGES MESSIEURS",
    "42": "400 m 4 NAGES DAMES",
    "92": "400 m 4 NAGES MESSIEURS",
    "46": "4 x 100 m 4 NAGES DAMES", 
    "96": "4 x 100 m 4 NAGES MESSIEURS",
    "47": "4 x 50 m NAGE LIBRES DAMES",
    "97": "4 x 50 m NAGE LIBRES MESSIEURS",
    "48": "4 x 50 m 4 NAGES DAMES",
    "98": "4 x 50 m 4 NAGES MESSIEURS",
    "37": "4 x 50 m 4 NAGES Mixte",
    "87": "4 x 50 NAGE LIBRES Mixte"
}

event_minimas = {
    "01": {  # 50 NL
        "Poussin": "0:45.08",
        "Benjamin": "0:43.49",
        "Minime": "0:34.89",
        "Cadet": "0:33.25",
        "18+": "0:30.71"
    },
    "51": {  # 50 NL
        "Poussin": "0:42.32",
        "Benjamin": "0:38.23",
        "Minime": "0:30.82",
        "Cadet": "0:29.01",
        "18+": "0:26.34"
    },
    "02": {  # 100 NL
        "Poussin": "1:37.40",
        "Benjamin": "1:30.89",
        "Minime": "1:13.42",
        "Cadet": "1:11.01",
        "18+": "1:06.45"
    },
    "52": {  # 100 NL
        "Poussin": "1:29.56",
        "Benjamin": "1:22.11",
        "Minime": "1:07.03",
        "Cadet": "1:02.56",
        "18+": "0:58.06"
    },
    "03": {  # 200 NL
        
        "Benjamin": "2:50.00",
        "Minime": "2:39.84",
        "Cadet": "2:34.46",
        "18+": "2:23.26"
    },
    "53": {  # 200 NL
        "Benjamin": "3:01.87",
        "Minime": "2:21.02",
        "Cadet": "2:14.46",
        "18+": "2:07.34"
    },
    "04": {  # 400 NL
        "Poussin": "7:05.23",
        "Benjamin": "6:43.43",
        
        "Minime": "5:39.94",
        "Cadet": "5:22.98",
        "18+": "5:00.18"
    },
    "54": {  # 400 NL
        "Poussin": "6:31.50",
        "Benjamin": "6:14.53",
        "Minime": "5:01.29",
        "Cadet": "4:46.20",
        "18+": "4:30.50"
    },
    "05": {  # 800 NL
        "Benjamin": "13:40.33",
        "Minime": "11:25.38",
        "Cadet": "10:54.71",
        "18+": "10:15.64"
    },
    "55": {  # 800 NL
        "Benjamin": "12:46.93",
        "Minime": "10:30.86",
        "Cadet": "9:58.40",
        "18+": "9:21.00"
    },
    "06": {  # 1500 NL
        "Minime": "21:54.58",
        "Cadet": "21:00.30",
        "18+": "19:50.30"
    },
    "56": {  # 1500 NL
        "Minime": "20:45.85",
        "Cadet": "18:47.64",
        "18+": "17:45.61"
    },
    "11": {  # 50 Dos
        "Poussin": "0:52.09",
        "Benjamin": "0:50.07",
        "Minime": "0:42.78",
        "Cadet": "0:38.13",
        "18+": "0:35.95"
    },
    "61": {  # 50 Dos
        "Poussin": "0:48.87",
        "Benjamin": "0:46.01",
        "Minime": "0:37.38",
        "Cadet": "0:33.06",
        "18+": "0:30.61"
        
    },
    "12": {  # 100 Dos
        "Poussin": "1:50.23",
        "Benjamin": "1:44.61",
        "Minime": "1:27.72",
        "Cadet": "1:21.73",
        "18+": "1:16.30"
    },
    "62": {  # 100 Dos
        "Poussin": "1:45.52",
        "Benjamin": "1:36.97",
        "Minime": "1:19.60",
        "Cadet": "1:13.11",
        "18+": "1:06.38"
    },
    "13": {  # 200 Dos
        "Benjamin": "3:37.58",
        "Minime": "3:02.55",
        "Cadet": "2:50.30",
        "18+": "2:42.40"
    },
    "63": {  # 200 Dos
        "Benjamin": "3:23.81",
        "Minime": "2:46.02",
        "Cadet": "2:35.84",
        "18+": "2:24.53"
    },
    "21": {  # 50 Brasse
        "Poussin": "0:59.10",
        "Benjamin": "0:56.09",
        "Minime": "0:44.49",
        "Cadet": "0:41.20",
        "18+": "0:38.68"
    },
    "71": {  # 50 Brasse
        "Poussin": "0:54.43",
        "Benjamin": "0:50.20",
        "Minime": "0:40.43",
        "Cadet": "0:36.96",
        "18+": "0:33.73"
        
    },
    "22": {  # 100 Brasse
        "Poussin": "2:01.07",
        "Benjamin": "1:57.76",
        "Minime": "1:36.23",
        "Cadet": "1:34.19",
        "18+": "1:25.04"
        
    },
    "72": {  # 100 Brasse
        "Poussin": "1:57.41",
        "Benjamin": "1:46.31",
        "Minime": "1:27.39",
        "Cadet": "1:19.90",
        "18+": "1:14.47"
    },
    "23": {  # 200 Brasse
        "Benjamin": "4:05.35",
        "Minime": "3:25.55",
        "Cadet": "3:20.58",
        "18+": "3:02.57"
    },
    "73": {  # 200 Brasse
        "Benjamin": "3:42.42",
        "Minime": "3:06.04",
        "Cadet": "2:50.45",
        "18+": "2:40.01"
        
    },
    "31": {  # 50 Papillon
        "Poussin": "0:52.28",
        "Benjamin": "0:50.08",
        "Minime": "0:38.14",
        "Cadet": "0:36.94",
        "18+": "0:34.17"
    },
    "81": {  # 50 Papillon
        "Poussin": "0:47.25",
        "Benjamin": "0:46.10",
        "Minime": "0:32.88",
        "Cadet": "0:30.45",
        "18+": "0:28.25"
        
    },
    "32": {  # 100 Papillon
        "Poussin": "1:59.33",
        "Benjamin": "1:45.93",
        "Minime": "1:25.74",
        "Cadet": "1:20.22",
        "18+": "1:14.11"
    },
    "82": {  # 100 Papillon
        "Poussin": "1:44.49",
        "Benjamin": "1:38.21",
        "Minime": "1:15.37",
        "Cadet": "1:07.40",
        "18+": "1:04.75"
    },
    "33": {  # 200 Papillon
        "Benjamin": "3:45.00",
        "Minime": "3:10.76",
        "Cadet": "2:52.54",
        "18+": "2:42.65"
    },
    "83": {  # 200 Papillon
        "Benjamin": "3:26.00",
        "Minime": "2:48.25",
        "Cadet": "2:28.70",
        "18+": "2:22.18"
    },
    "41": {  # 200 4 Nages
        "Poussin": "3:54.03",
        "Benjamin": "3:30.45",
        "Minime": "3:00.71",
        "Cadet": "2:53.09",
        "18+": "2:42.50"
    },
    "91": {  # 200 4 Nages
        "Poussin": "3:39.95",
        "Benjamin": "3:17.86",
        "Minime": "2:43.44",
        "Cadet": "2:33.34",
        "18+": "2:25.57"
        
    },
    "42": {  # 400 4 Nages
        
        "Minime": "6:29.77",
        "Cadet": "5:59.18",
        "18+": "5:42.86"
    },
    "92": {  # 400 4 Nages
        "Minime": "5:37.25",
        "Cadet": "5:27.13",
        "18+": "5:10.61"
    },
    "47": {  # 4 x 50 m NAGE LIBRES DAMES
        "Poussin": "3:12.32",
        "Benjamin": "2:54.08"
    },
    "97": {  # 4 x 50 m NAGE LIBRES DAMES
        "Benjamin": "2:54.08"
    },
    "48": {  # 4 x 50 m 4 NAGES DAMES
        "Poussin": "3:36.61",
    },
    "98": {  # 4 x 50 m 4 NAGES DAMES
        "Benjamin": "3:18.88"
    },
    "43": {  # 4 x 100 m NAGE LIBRE DAMES
        
        "Minime": "4:53.68",
        "Cadet": "4:44.04",
        "18+": "4:25.80"
    },
    "46": {  # 4 x 100 m 4 NAGES DAMES
        
        "Minime": "5:43.11",
        "Cadet": "5:27.15",
        "18+": "5:01.90"
    },
    "93": {  # 4 x 100 m NAGE LIBRE MESSIEURS
        "Minime": "4:28.12",
        "Cadet": "4:10.24",
        "18+": "3:52.24"
    },
    "96": {  # 4 x 100 m 4 NAGES MESSIEURS
        "Minime": "5:09.39",
        "Cadet": "4:42.97",
        "18+": "4:21.94"
    },
    "44": {  # 4 x 200 m NAGE LIBRE DAMES
        
        "Minime": "10:29.36",
        "Cadet": "10:17.84",
        "18+": "9:33.04"
    },
    "94": {  # 4 x 200 m NAGE LIBRE MESSIEURS
        "Minime": "9:24.08",
        "Cadet": "8:57.84",
        "18+": "8:29.36"
    }
}

classement = {
    "Championnat d'Hiver 2024 MCJS18+": {
        "Poussin": {
            # "Club A": 120,
            # "Club B": 110,
            # "Club C": 95
        },
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        },
        "Minime": {
            # "Club A": 140,
            # "Club B": 135,
            # "Club C": 120
        },
        "Cadet": {
            # "Club A": 150,
            # "Club B": 145,
            # "Club C": 130
        },
        "18+": {
            # "Club A": 160,
            # "Club B": 155,
            # "Club C": 140
        }
    },
    "Championnat d'Été 2024 MCJS18+": {
        "Poussin": {
            # "Club A": 120,
            # "Club B": 110,
            # "Club C": 95
        },
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        },
        "Minime": {
            # "Club A": 140,
            # "Club B": 135,
            # "Club C": 120
        },
        "Cadet": {
            # "Club A": 150,
            # "Club B": 145,
            # "Club C": 130
        },
        "18+": {
            # "Club A": 160,
            # "Club B": 155,
            # "Club C": 140
        },
        "TC": {
            # "Club A": 160,
            # "Club B": 155,
            # "Club C": 140
        }
    },
    "Championnat d'Été 2023 MCJS18+": {
        "Poussin": {
            # "Club A": 120,
            # "Club B": 110,
            # "Club C": 95
        },
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        },
        "Minime": {
            # "Club A": 140,
            # "Club B": 135,
            # "Club C": 120
        },
        "Cadet": {
            # "Club A": 150,
            # "Club B": 145,
            # "Club C": 130
        },
        "18+": {
            # "Club A": 160,
            # "Club B": 155,
            # "Club C": 140
        }
    },
    "Championnat d'Hiver 2024 Benjamins": {
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        }
    },
    "Championnat d'Été 2024 Benjamins": {
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        }
    },
    "Championnat d'Hiver 2023 Benjamins": {
        "Poussin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        },
        "Benjamin": {
            # "Club A": 130,
            # "Club B": 125,
            # "Club C": 100
        }
    },
}

events_data = {
    championnat: {
        event_code: {
            category: [] for category in classement[championnat].keys()
        } for event_code in event_names.keys()
    } for championnat in classement.keys()
}

swimmers = [
            {
                "Name": "",
                "Club": "",
                "Events": [
                    {
                        "Event": "",
                        "Results": [
                            {
                                "Championnat": "",
                                "Time": "",
                                "Place": "",
                                "Points": ""
                            }
                        ]
                    }
                ],
                "Category": "",
                "Year of Birth": "" 
        }
    ]


# Example data structure for events_data
# events_data = {
#     1: {  # Soup index
#         "Championnat d'Hiver 2024 MCJS18+": {  # Event code
#             "Poussin": [  # Category
#                 {'Event': '50 m NAGE LIBRE DAMES', 'Name': 'Swimmer Name', 'Place': 1, 'Year of Birth': '2008', 'Club': 'Club A', 'Time': '00:25.00', 'Category': 'Poussin', 'Points': 100},
#                 {'Event': '50 m NAGE LIBRE DAMES', 'Name': 'Another Swimmer', 'Place': 2, 'Year of Birth': '2009', 'Club': 'Club B', 'Time': '00:26.00', 'Category': 'Poussin', 'Points': 90}
#             ],
#             "Benjamin": [  # Another category
#                 {'Event': '50 m NAGE LIBRE DAMES', 'Name': 'Swimmer Name', 'Place': 1, 'Year of Birth': '2006', 'Club': 'Club A', 'Time': '00:24.00', 'Category': 'Benjamin', 'Points': 100},
#                 {'Event': '50 m NAGE LIBRE DAMES', 'Name': 'Another Swimmer', 'Place': 2, 'Year of Birth': '2007', 'Club': 'Club B', 'Time': '00:25.00', 'Category': 'Benjamin', 'Points': 90}
#             ]
#         }
#     }
# }




def print_event_data(event_code,event_data):
    if event_code in events_data[event_data]:
        print(f"Event: {event_names[event_code]}")
        for category, data in events_data[event_data][event_code].items():
            if not data:  # Check if the category is empty
                continue
            print("--------------------------")
            print(f"Category: {category}")
            print("--------------------------")
            for swimmer_data in data:
                print(f"{swimmer_data['Place']}. {swimmer_data['Name']}  ({swimmer_data['Year of Birth']}) Time: {swimmer_data['Time']} Points: {swimmer_data['Points']} Club: {swimmer_data['Club']}")
            print()
    else:
        print(f"Event code {event_code} not found.")


def check_ranking(category, rank, notCounted,event):
    if debug_mode:
        print("Calculating rank :",rank, " in :", category)
    if "x" in event_names[event] and category == "Poussin":
        return rank <= 8
    if notCounted != 0 :
        if debug_mode:
            print("There's", notCounted, "international swimmer ranked in this event")
    if category == "Poussin":
        return rank <= 32 + notCounted
    elif category == "Benjamin":
        return rank <=24 + notCounted
    elif category == "Minime":
        return rank <= 16 + notCounted
    elif category == "Cadet" or category == "18+":
        return rank <= 8 + notCounted
    else:
        return False


# Function to compare swimmer times to minimas
def time_to_seconds(time_str):
    try:
        if ':' in time_str:
            minutes, rest = time_str.split(':')
            seconds, milliseconds = rest.split('.')
            total_seconds = int(minutes) * 60 + int(seconds) + int(milliseconds) / 100
        else:
            seconds, milliseconds = time_str.split('.')
            total_seconds = int(seconds) + int(milliseconds) / 100
        return total_seconds
    except ValueError:
        # print(f"Invalid time format: {time_str}",)
        return 999999

def compare_times(swimmer_time, minima_time):
    swimmer_seconds = time_to_seconds(swimmer_time)
    minima_seconds = time_to_seconds(minima_time)
    
    if swimmer_seconds is None or minima_seconds is None:
        return False
    
    return swimmer_seconds <= minima_seconds

# Function to categorize swimmers by year of birth
def categorize_swimmer(year_of_birth, championnat):
    year_to_date = datetime.now().year
    if year_of_birth != 'Naissance' :
        year_of_birth = int(year_of_birth)
        if "TC" in championnat:
            return "18+"
        if "2023" in championnat:
            year_to_date -= 1
            print("NOOOOOWWWW")
        if year_to_date - year_of_birth in [10, 11]:
            return "Poussin"
        elif year_to_date - year_of_birth in [12, 13]:
            return "Benjamin"
        elif year_to_date - year_of_birth in [14, 15]:
            return "Minime"
        elif year_to_date - year_of_birth in [16, 17]:
            return "Cadet"
        else:
            return "18+"

# Loop through the elements to find swimmers' data within each event
pointsFound = 1
championnat_id = 0
for soup in soups:
    championnat = list(classement.keys())[championnat_id]
    championnat_id += 1
    for element in soup.find_all():
        
        if element.name == 'a' and element.get('name') and element.get('name') != "top": # Get every a tag with an attribute name
            # Start a new event
            current_event = element.get('name') # That attribute is the event code name
            events_data[championnat][current_event] = {category: [] for category in classement.keys()}
            if debug_mode:
                print('__________________________________________________________________')
                print("|")
                print("|     --->     A new event is started : (", current_event, ")", event_names[current_event])
                print('|__________________________________________________________________')
            notCounted = 0
            lastCategory = ""
        elif current_event and element.name == 'tr':  # Assuming the data is within table rows of the previously found a tag with an attribute name of theevent code
            columns = element.find_all('td')
            if columns[0].text.strip() == "Place":
                continue
            if len(columns) > 5:  # Adjust based on actual number of columns
                print(current_event)
                if "x" not in event_names[current_event]:
                    name = columns[1].text.strip()
                    club = columns[4].text.strip()  # Corrected index for club
                    year_of_birth = columns[3].text.strip()  # Corrected index for year of birth
                    time = columns[5].text.strip()  # Corrected index for time
                    nat = columns[2].text.strip()
                    category = categorize_swimmer(year_of_birth, championnat)
                    try:
                        place = int(columns[0].text.strip().rstrip('.'))
                    except ValueError:
                        # print(columns[0].text.strip())
                        place = 0
                    try:
                        points = int(columns[6].text.strip())
                    except ValueError:
                        points = 0 
                    if debug_mode:
                        print("This is indivdual:", club, "points:", points, "rank:", place, )
                else:  # Relay events
                    if pointsFound == 1:
                        year_of_birth = columns[3].text.strip()
                        category = categorize_swimmer(year_of_birth,championnat)
                        try:
                            place = int(columns[0].text.strip().rstrip('.'))
                        except ValueError:
                            # print(columns[0].text.strip())
                            place = 0
                        name = columns[1].text.strip()
                        # print(pointsFound, name)
                        pointsFound += 1
                        continue
                    elif pointsFound <= 4:
                        name += "," + columns[1].text.strip()  # Corrected index for name
                        # print(pointsFound, columns[1].text.strip())
                        if pointsFound == 4:
                            club = columns[4].text.strip()
                            time = columns[5].text.strip()
                            try:
                                points = int(columns[6].text.strip())
                            except ValueError:
                                points = 0
                            pointsFound = 1
                        else :
                            pointsFound += 1
                            continue
                        
                    if debug_mode:
                        print("This is relay:", club, "points:", points, "rank:", place)
                pointsFound = 1
                
                swimmer_data = {
                    'Event': event_names.get(current_event, f"Event {current_event}"),
                    'Championnat' : championnat,
                    'Name': name.replace('\xa0', ' '),
                    'Place': place,
                    'Year of Birth': year_of_birth,
                    'Club': club,
                    'Time': time,
                    'Category': category,
                    'Points' : points,
                    'Nationality' : nat
                }

                events_data[championnat][current_event].setdefault(category, []).append(swimmer_data) 
                if debug_mode:
                    print(events_data[championnat][current_event][category])
                    print(swimmer_data)
                    
                # Check if points are an integer
                
                minima = event_minimas.get(current_event, {}).get(category, "N/A")
                # print(minima)
                # Number of swimmers that does NOT count
                if (nat != 'TUN' and place in range(1, 9 + notCounted)) or (club == 'LP'):
                    lastCategory = category
                    notCounted += 1 
                    # print(category, place, name, club, year_of_birth, nat, points)
                    continue
                if lastCategory != category: 
                    notCounted = 0
                    
                if debug_mode:
                    print("time : ", time,"minima : ",minima)
                if compare_times(time,minima) :
                    # print("Time is: ",time,"| Minimas is: ", minima)
                    if check_ranking(category,place,notCounted,current_event):
                        if debug_mode:
                            print("++++++++++ SWIMMER IN ++++++++++")
                        if club in classement[championnat][category]:
                            if current_event in ["37","87","43","93","44","94","46","96","47","97","48","98"]:
                                points *= 2
                            classement[championnat][category][club] += points
                        else:
                            classement[championnat][category][club] = points
                    else :
                        if debug_mode:
                            print("---------- SWIMMER OUT ----------")
                if debug_mode:
                    print(event_names[current_event], category, place, name, club, category, nat, points)
                    print()

def createExcel():
    # Create a new Excel file named 'swimmers_results.xlsx'
    with pd.ExcelWriter('swimmers_results.xlsx', engine='openpyxl') as writer:
        # Prepare a DataFrame to hold all swimmer data
        all_swimmer_data = []

        # Iterate through each swimmer in the swimmers list
        for swimmer in swimmers:
            for event in swimmer['Events']:
                all_swimmer_data.append({
                    "Events": event['Event'],
                    "Championnat": event['Championnat'],
                    "Category": swimmer['Category'],
                    "Year of Birth": swimmer['Year of Birth'],
                    "Club": swimmer['Club'],
                    "Time": event['Time'],
                    "Place": event['Place'],
                })

        # Convert the list to a DataFrame
        df = pd.DataFrame(all_swimmer_data)

        # Write the DataFrame to the Excel file
        df.to_excel(writer, index=False, header=True)




# Sort the classement dictionary by points in descending order
sorted_classement = {championship: {category: dict(sorted(clubs.items(), key=lambda x: x[1], reverse=True)) for category, clubs in categories.items()} for championship, categories in classement.items()}

def find_swimmer_with_most_medals_in_championship(championship):
    """
    This function finds the swimmer who collected the most medals across all events and categories within a given championship.
    
    Parameters:
    - championship (str): The name of the championship to consider.
    
    Returns:
    - A tuple containing the name of the swimmer with the most medals, their club, and the total medals.
    """
    swimmers_medals = {}
    
    # Check if the championship exists in the events data
    if championship in events_data:
        # Iterate through each event in the championship
        for event_code, categories in events_data[championship].items():
            # Iterate through each category in the event
            for category, data in categories.items():
                # Iterate through each swimmer's data in the category
                for swimmer_data in data:
                    # Check if the swimmer's place is in the top 3
                    if swimmer_data['Place'] <= 3:
                        # Add the swimmer's medal to their total medals
                        if swimmer_data['Name'] not in swimmers_medals:
                            swimmers_medals[swimmer_data['Name']] = {'Club': swimmer_data['Club'], 'Medals': 0}
                        swimmers_medals[swimmer_data['Name']]['Medals'] += 1
    
    most_medals_swimmer = max(swimmers_medals, key=lambda x: swimmers_medals[x]['Medals'])
    most_medals = swimmers_medals[most_medals_swimmer]['Medals']
    most_medals_club = swimmers_medals[most_medals_swimmer]['Club']
    
    # Return the swimmer with the most medals, their club, and the total medals
    print(f"Swimmer with the most medals in championship {championship}: {most_medals_swimmer} from {most_medals_club} with {most_medals} medals")


def calculate_time_progression(event_data):
    """
    This function calculates the progression between times in a single event for a swimmer.
    
    Parameters:
    - event_data (list): The list of swimmer data for a single event.
    
    Returns:
    - A list of time progressions for the swimmer in the event.
    """
    time_progression = []
    previous_time = None
    for swimmer_data in event_data:
        current_time = swimmer_data['Time']
        if previous_time is not None:
            time_progression.append(time_to_seconds(current_time) - time_to_seconds(previous_time))
        previous_time = current_time
    return time_progression


def find_swimmer_with_most_points_in_championship(championship):
    """
    This function finds the swimmer who collected the most points across all events and categories within a given championship.
    
    Parameters:
    - championship (str): The name of the championship to consider.
    
    Returns:
    - A tuple containing the name of the swimmer with the most points, their club, and the total points.
    """
    # Demonstration of swimmers_points data structure
    # swimmers_points = {
    #     'Swimmer Name': {'Club': 'Club A', 'Points': 100},
    #     'Another Swimmer': {'Club': 'Club B', 'Points': 90}
    # }
    swimmers_points = {}
    # Check if the championship exists in the events data
    if championship in events_data:
        # Iterate through each event in the championship
        for event_code, categories in events_data[championship].items():
            # Iterate through each category in the event
            for category, data in categories.items():
                # Iterate through each swimmer's data in the category
                for swimmer_data in data:
                    # Add the swimmer's points to their total points
                    if "," not in swimmer_data['Name'] :    
                        if swimmer_data['Name'] not in swimmers_points:
                            print(swimmer_data['Name'],swimmer_data['Club'])
                            swimmers_points[swimmer_data['Name']] = {'Club': swimmer_data['Club'], 'Points': 0}
                        swimmers_points[swimmer_data['Name']]['Points'] += swimmer_data['Points']
    
    most_points_swimmer = max(swimmers_points, key=lambda x: swimmers_points[x]['Points'])
    most_points = swimmers_points[most_points_swimmer]['Points']
    most_points_club = swimmers_points[most_points_swimmer]['Club']
    
    # Return the swimmer with the most points, their club, and the total points
    print(f"Swimmer with the most points: {most_points_swimmer} from {most_points_club} with {most_points} points")

    # Return the swimmer with the most points, their club, and the total points
    print()


def collect_swimmers_names() :
    for championnat, events in events_data.items():
        for event_code, categories in events.items():
            for category, data in categories.items():
                for swimmer in data:
                    if swimmer['Club'] == "CA" and "," not in swimmer['Name']:
                        if swimmer['Name'] not in [swimmer_info['Name'] for swimmer_info in swimmers]:
                            swimmers.append({
                            "Name": swimmer['Name'],
                            "Club": swimmer['Club'],
                            "Events": [
                                {
                                    "Event": swimmer['Event'],
                                    "Results": [
                                        {
                                            "Championnat": championnat,
                                            "Time": swimmer['Time'],
                                            "Place": swimmer['Place'],
                                            "Points": swimmer['Points']
                                        }
                                    ]
                                }
                            ],
                            "Category": swimmer['Category'],
                                "Year of Birth": swimmer['Year of Birth']
                            })
                        else: 
                            # This block of code is executed when the swimmer is already in the swimmers list.
                            # It iterates through the swimmers list to find the swimmer with the matching name.
                            # Once found, it appends the new event result to the swimmer's 'Events' list.
                            for swimmer_info in swimmers:
                                if swimmer_info['Name'] == swimmer['Name']:
                                    event_exists = False
                                    for event in swimmer_info['Events']:
                                        if event['Event'] == event_names[event_code]:
                                            event['Results'].append({
                                                "Championnat": championnat,
                                                "Time": swimmer['Time'],
                                                "Place": swimmer['Place'],
                                                "Points": swimmer['Points']
                                            })
                                            event_exists = True
                                            break
                                    if not event_exists:
                                        swimmer_info['Events'].append({
                                            "Event": event_names[event_code],
                                            "Results": [
                                                {
                                                    "Championnat": championnat,
                                                    "Time": swimmer['Time'],
                                                    "Place": swimmer['Place'],
                                                    "Points": swimmer['Points']
                                                }
                                            ]
                                        })
def copy_swimmers_to_excel():
    # Create a new Excel file named 'swimmers_sheets' using openpyxl engine, creating the file if it doesn't exist
    with pd.ExcelWriter('swimmers_sheets_CA.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Iterate through each swimmer in the swimmers list
        for swimmer in swimmers:
            swimmer_data_list = []
            if not swimmer['Name'] or "," in swimmer['Name']:
                print("NAME EMPTY or Relay")
                continue

            # Prepare data for the swimmer
            swimmer_data = {
                "Name": swimmer['Name'],
                "Club": swimmer['Club'],                
                "Year of birth": swimmer['Year of Birth'],
                "Events": []
            }
            
            # Collect event results
            for event in swimmer['Events']:
                event_results = {
                    "Event": event['Event'],
                    "Results": []
                }
                # Find the lowest time for highlighting using the time_to_seconds function

                lowest_time = min(time_to_seconds(result['Time']) for result in event['Results'])
                if debug_mode:
                    print(lowest_time, event['Event'])
                
                for result in event['Results']:
                    result_row = {
                        "Championship": result['Championnat'],
                        "Time": result['Time'],
                        "Place": result['Place'],
                        "Points": result['Points'],
                        "Highlight": False
                    }
                    
                    # Highlight the lowest time
                    if time_to_seconds(result['Time']) == lowest_time:
                        result_row["Highlight"] = True  # Mark for highlighting
                        if debug_mode:
                            print(result['Time'], lowest_time, result_row['Highlight'])

                    event_results["Results"].append(result_row)

                swimmer_data["Events"].append(event_results)

            # Convert swimmer data into a DataFrame
            for event in swimmer_data["Events"]:
                for result in event["Results"]:
                    swimmer_data_list.append({
                        'Event': event['Event'],
                        'Championship': result['Championship'],
                        'Time': result['Time'],
                        'Place': result['Place'],
                        'Points': result['Points']
                    })
            swimmer_df = pd.DataFrame(swimmer_data_list)

            # Write the DataFrame to the Excel file, specifying the sheet name and not including the index
            swimmer_df.to_excel(writer, sheet_name=swimmer['Name'], index=False, header=True)

            # Highlight the lowest time in green
            workbook = writer.book
            worksheet = workbook[swimmer['Name']]
            # for row in range(2, len(swimmer_df) + 2):  # Start from row 2 to skip header
            #     if swimmer_df.loc[row - 2, 'Highlight']:
            #         print(row, swimmer_df.loc[row - 2, 'Highlight'])
            #         worksheet.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill

        # Write the summary DataFrame to the Excel file, specifying the sheet name and not including the index
        summary_list = []
        for swimmer in swimmers:
            if not swimmer['Name'] or "," in swimmer['Name']:
                continue 
            summary_list.append({
                "Name": f'=HYPERLINK("[#\'{swimmer["Name"]}\'!A1]", "{swimmer["Name"]}")',
                "Club": swimmer['Club'],                
                "Category": swimmer['Category'],
                "Year of Birth": swimmer['Year of Birth']
            })

        summary_df = pd.DataFrame(summary_list)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=True)



def printClassement() :
    print()
    print("------------------------------")
    print("----------Classement----------")
    print("------------------------------")
    print()
    i = 1
    for selected_championnat in classement.keys():
        print(i, selected_championnat)
        i+=1
    selected_index = int(input("Enter your choice:"))
    selected_championnat = list(classement.keys())[selected_index - 1]
    if selected_championnat in sorted_classement:
        categories = sorted_classement[selected_championnat]
        for category, clubs in categories.items():
            if not clubs:
                continue
            print(f"Catégorie: {category}")
            for club, points in clubs.items():
                print(f"{club}: {points} points")
            print()
                

print('__________________________________________________________________')
print("|")
print("|     Sources:")
for championnat in classement.keys():
    print("|     ",championnat)
print('|__________________________________________________________________') 
while True:
    # Display menu options
    print("Menu:")
    print("1. Classement des clubs")
    print("2. Resultat par epreuve")
    print("3. Generate Excel sheets")
    print("Exit [Enter]")
    print()
    print("Testing options:")
    print("4. find_swimmer_with_most_points_in_event()")
    print("5. find_swimmer_with_most_medals_in_championship()")
    print("6. collect_swimmers_names()")

    choice = input("Enter your choice: ")

    if choice == "1":
        # Code for Option 1
        printClassement()
    elif choice == "2":
        # Code for Option 2
        print("Select an event to view results:")   
        for event_code, event_name in event_names.items():
            print(f"{event_code}: {event_name}")
        selected_event = str(input("Enter the event code: "))
        if selected_event in event_names:
            print(f"Results for {event_names[selected_event]}:")
            i = 1
            for selected_championnat in classement.keys():
                print(i, selected_championnat)
                i+=1
            print("championnat: ")
            selected_index = int(input("Enter your choice:"))
            selected_championnat = list(classement.keys())[selected_index - 1]
            # Assuming there's a function to display event results
            print_event_data(selected_event,selected_championnat)
        else:
            print("Invalid event code. Please try again.")
        
    elif choice == "3":
        # Code for Option 3
        createExcel()
    elif choice == "4":
            i = 1
            for selected_championnat in classement.keys():
                print(i, selected_championnat)
                i+=1
            print("championnat: ")
            selected_index = int(input("Enter your choice:"))
            selected_championnat = list(classement.keys())[selected_index - 1]
            find_swimmer_with_most_points_in_championship(selected_championnat)
    elif choice == "5":
            i = 1
            for selected_championnat in classement.keys():
                print(i, selected_championnat)
                i+=1
            print("championnat: ")
            selected_index = int(input("Enter your choice:"))
            selected_championnat = list(classement.keys())[selected_index - 1]
            find_swimmer_with_most_medals_in_championship(selected_championnat)
    elif choice == "6":
        collect_swimmers_names()
        copy_swimmers_to_excel()
    
        print("Swimmers' Information:")
        for swimmer in swimmers:
            print(f"Name: {swimmer['Name']}, Club: {swimmer['Club']}, Category: {swimmer['Category']}, Year of Birth: {swimmer['Year of Birth']}")
            for event in swimmer['Events']:
                print(f"Event: {event['Event']}")
                for result in event['Results']:
                    print(f"Championship: {result['Championnat']}, Place: {result['Place']}, Time: {result['Time']}, Points: {result['Points']}")
    else:
        print("Exiting...")
        break